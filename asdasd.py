import csv
import re
from datetime import datetime
import os
from statistics import mean
import openpyxl
from matplotlib import ticker
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit


def quick_quit(msg):
    print(msg)
    exit(0)


currency_to_rub = {
    "Манаты": 35.68,
    "Белорусские рубли": 23.91,
    "Евро": 59.90,
    "Грузинский лари": 21.74,
    "Киргизский сом": 0.76,
    "Тенге": 0.13,
    "Рубли": 1,
    "Гривны": 1.64,
    "Доллары": 60.66,
    "Узбекский сум": 0.0055,
}

currency = {
    "AZN": "Манаты",
    "BYR": "Белорусские рубли",
    "EUR": "Евро",
    "GEL": "Грузинский лари",
    "KGS": "Киргизский сом",
    "KZT": "Тенге",
    "RUR": "Рубли",
    "UAH": "Гривны",
    "USD": "Доллары",
    "UZS": "Узбекский сум",
}


class UserInput:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')


class DataSet:
    def __init__(self, file_name):
        if os.stat(file_name).st_size == 0:
            quick_quit("Пустой файл")

        self.data = [row for row in csv.reader(open(file_name, encoding="utf_8_sig"))]
        self.names = self.data[0]
        self.all_data = [row for row in self.data[1:] if len(row) == len(self.names) and row.count('') == 0]

        if len(self.all_data) == 0:
            quick_quit('Нет данных')


class StaticInfo:
    salaries_by_year: dict = {}
    vacancies_by_year: dict = {}
    inp_vacancy_salary: dict = {}
    inp_vacancy_count: dict = {}
    salaries_areas: dict = {}
    vacancies_areas: dict = {}
    vacancies: int = 0
    dicts_list_by_year: list = []
    dicts_list_by_area: list = []

    def __init__(self):
        for i in range(2007, 2023):
            self.salaries_by_year[i] = []
            self.vacancies_by_year[i] = 0
            self.inp_vacancy_salary[i] = []
            self.inp_vacancy_count[i] = 0

    @staticmethod
    def sort_dict(dictionary):
        sorted_tuples = sorted(dictionary.items(), key=lambda item: item[1], reverse=True)[:10]
        sorted_dict = {k: v for k, v in sorted_tuples}
        return sorted_dict

    @staticmethod
    def check_len_dic(dic, vac_name=""):
        new_dic = {}
        for name, item in dic.items():
            if len(item) != 0:
                new_dic[name] = int(mean(item))
        if vac_name != "" and len(new_dic) == 0:
            return {2022: 0}
        return new_dic

    @staticmethod
    def check_int_dic(dic, vac_name=""):
        new_dic = {}
        for name, item in dic.items():
            if item != 0:
                new_dic[name] = item
        if vac_name != "" and len(new_dic) == 0:
            return {2022: 0}
        return new_dic

    def check_one_percent(self):
        vacancies_areas_dict = {}
        salaries_areas_dict = {}
        self.vac_with_others = {"Другие": 0}
        for key, value in self.vacancies_areas.items():
            percent = value / self.vacancies
            if percent >= 0.01:
                salaries_areas_dict[key] = int(mean(self.salaries_areas[key]))
                vacancies_areas_dict[key] = round(percent, 4)
            else:
                self.vac_with_others["Другие"] += percent
        self.salaries_areas = salaries_areas_dict
        self.vacancies_areas = vacancies_areas_dict

    def print_result(self, vac_name):
        self.check_one_percent()
        ch_salaries_by_year = self.check_len_dic(self.salaries_by_year)
        ch_vacancies_by_year = self.check_int_dic(self.vacancies_by_year)
        ch_inp_vacancy_salary = self.check_len_dic(self.inp_vacancy_salary, vac_name)
        ch_inp_vacancy_count = self.check_int_dic(self.inp_vacancy_count, vac_name)
        ch_salaries_areas = self.sort_dict(self.salaries_areas)
        ch_vacancies_areas = self.sort_dict(self.vacancies_areas)

        print("Динамика уровня зарплат по годам:", ch_salaries_by_year)
        print("Динамика количества вакансий по годам:", ch_vacancies_by_year)
        print("Динамика уровня зарплат по годам для выбранной профессии:", ch_inp_vacancy_salary)
        print("Динамика количества вакансий по годам для выбранной профессии:", ch_inp_vacancy_count)
        print("Уровень зарплат по городам (в порядке убывания):", ch_salaries_areas)
        print("Доля вакансий по городам (в порядке убывания):", ch_vacancies_areas)
        self.dicts_list_by_year = [ch_salaries_by_year, ch_inp_vacancy_salary,
                                   ch_vacancies_by_year, ch_inp_vacancy_count]
        self.dicts_list_by_area = [ch_salaries_areas, ch_vacancies_areas]


class Vacancy:
    name: str
    salary_from: int or float
    salary_to: int or float
    salary_currency: str
    area_name: str
    published_at: str
    salary: int

    def __init__(self, pers_data):
        for name, item in pers_data.items():
            self.__setattr__(name, self.formatter(name, item))

        self.salary = int(currency_to_rub[self.salary_currency] * (
                float(self.salary_from.replace(' ', '')) + float(self.salary_to.replace(' ', ''))) // 2)

    @staticmethod
    def formatter(name, item):
        if name == 'salary_currency':
            return currency[item]
        elif name == "salary_to" or name == "salary_from":
            return '{:,}'.format(int(float(item))).replace(',', ' ')
        elif name == 'published_at':
            return int(datetime.strptime(item, '%Y-%m-%dT%H:%M:%S%z').strftime("%Y"))
        else:
            return item

    def get_vac_data(self, dicts, vac_name):
        year_key = self.published_at
        area_key = self.area_name

        dicts.vacancies_by_year[year_key] += 1
        dicts.salaries_by_year[year_key] += [self.salary]

        if area_key in dicts.vacancies_areas.keys():
            dicts.vacancies_areas[area_key] += 1
        else:
            dicts.vacancies_areas[area_key] = 1

        if area_key in dicts.salaries_areas.keys():
            dicts.salaries_areas[area_key] += [self.salary]
        else:
            dicts.salaries_areas[area_key] = [self.salary]

        if vac_name in self.name and vac_name != "":
            dicts.inp_vacancy_count[year_key] += 1
            dicts.inp_vacancy_salary[year_key] += [self.salary]

        dicts.vacancies += 1


class Report:
    rows_by_year = ["Год", "Средняя зарплата", "Средняя зарплата - ", "Количество вакансий",
                    "Количество вакансий - "]
    rows_by_area = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]

    def __init__(self, vac_name, dicts_by_year, dicts_by_area, vac_with_others):
        self.book = openpyxl.Workbook()
        self.book.remove(self.book.active)

        self.sheet_by_year = self.book.create_sheet("Статистика по годам")
        self.sheet_by_area = self.book.create_sheet("Статистика по городам")

        self.generate_excel(vac_name, dicts_by_year, dicts_by_area)
        self.generate_image(vac_name, dicts_by_year, dicts_by_area, vac_with_others)
        self.generate_pdf(vac_name, dicts_by_year, dicts_by_area)

    @staticmethod
    def generate_pdf(vac_name, dicts_by_year, dicts_by_area):
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        pdf_template = template.render(
            {'name': vac_name, 'by_year': dicts_by_year, 'by_area': dicts_by_area,
             'keys_0_area': list(dicts_by_area[0].keys()), 'values_0_area': list(dicts_by_area[0].values()),
             'keys_1_area': list(dicts_by_area[1].keys()), 'values_1_area': list(dicts_by_area[1].values())})

        options = {'enable-local-file-access': None}
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)

    @staticmethod
    def generate_image(vac_name, dicts_by_year, dicts_by_area, vac_with_others):
        y1_cities = np.arange(len(dicts_by_area[0].keys()))
        y1_cities_names = {}
        for key, value in dicts_by_area[0].items():
            if "-" in key or " " in key:
                key = key.replace("-", "-\n")
                key = key.replace(" ", "\n")
            y1_cities_names[key] = value

        x_nums = np.arange(len(dicts_by_year[0].keys()))
        width = 0.4
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2
        fig = plt.figure()

        ax = fig.add_subplot(221)
        ax.set_title("Уровень зарплат по годам")
        ax.bar(x_list1, dicts_by_year[0].values(), width, label="средняя з/п")
        ax.bar(x_list2, dicts_by_year[1].values(), width, label=f"з/п {vac_name.lower()}")
        ax.set_xticks(x_nums, dicts_by_year[0].keys(), rotation="vertical")
        ax.tick_params(axis="both", labelsize=8)
        ax.legend(fontsize=8)
        ax.grid(True, axis="y")

        ax = fig.add_subplot(222)
        ax.set_title("Количество вакансий по годам")
        ax.bar(x_list1, dicts_by_year[2].values(), width, label="Количество вакансий")
        ax.bar(x_list2, dicts_by_year[3].values(), width, label=f"Количество вакансий \n{vac_name.lower()}")
        ax.set_xticks(x_nums, dicts_by_year[2].keys(), rotation="vertical")
        ax.tick_params(axis="both", labelsize=8)
        ax.legend(fontsize=8)
        ax.grid(True, axis="y")

        ax = fig.add_subplot(223)
        ax.set_title("Уровень зарплат по городам")
        width = 0.8
        ax.barh(y1_cities, dicts_by_area[0].values(), width, align="center")
        ax.set_yticks(y1_cities, labels=y1_cities_names.keys(), horizontalalignment="right", verticalalignment="center")
        ax.tick_params(axis="x", labelsize=8)
        ax.tick_params(axis="y", labelsize=6)
        ax.xaxis.set_major_locator(ticker.MultipleLocator(40000))
        ax.invert_yaxis()
        ax.grid(True, axis="x")

        ax = fig.add_subplot(224)
        ax.set_title("Доля вакансий по городам")
        dicts_by_area[1]["Другие"] = vac_with_others["Другие"]
        ax.pie(dicts_by_area[1].values(), labels=dicts_by_area[1].keys(), textprops={'size': 6})
        ax.axis('equal')

        plt.tight_layout()
        plt.savefig("graph.png")

    def generate_excel(self, vac_name, dicts_by_year, dicts_by_area):
        thins = Side(border_style="thin", color="000000")

        self.set_value_first_sheet(vac_name, dicts_by_year, thins)
        self.set_value_second_sheet(dicts_by_area, thins)
        self.set_value(thins)

        self.book.save("report.xlsx")

    def set_value_first_sheet(self, vac_name, dicts_by_year, thins):
        # Установка столбцов первого листа
        for i, value in enumerate(self.rows_by_year, 1):
            self.sheet_by_year.cell(row=1, column=i).font = Font(bold=True)
            self.sheet_by_year.cell(row=1, column=i).border = Border(top=thins, bottom=thins, left=thins, right=thins)

            self.sheet_by_year.cell(row=1, column=i).value = value + vac_name if " - " in value else value

        # Установка значений первого листа
        for year, value in dicts_by_year[0].items():
            self.sheet_by_year.append(
                [year, value, dicts_by_year[1][year], dicts_by_year[2][year], dicts_by_year[3][year]])

    def set_value_second_sheet(self, dicts_by_area, thins):
        # Установка столбцов второго листа
        for i, value in enumerate(self.rows_by_area, 1):
            if value != "":
                self.sheet_by_area.cell(row=1, column=i).font = Font(bold=True)
                self.sheet_by_area.cell(row=1, column=i).border = Border(top=thins, bottom=thins, left=thins,
                                                                         right=thins)

            self.sheet_by_area.cell(row=1, column=i).value = value

        # Установка значений второго листа
        for i in range(len(dicts_by_area[0])):
            self.sheet_by_area.append([list(dicts_by_area[0].keys())[i],
                                       list(dicts_by_area[0].values())[i], "",
                                       list(dicts_by_area[1].keys())[i],
                                       list(dicts_by_area[1].values())[i]])

    def set_value(self, thins):
        dims = {}
        for row in self.sheet_by_year.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            self.sheet_by_year.column_dimensions[col].width = value + 2

        dims = {}
        for row in self.sheet_by_area.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            self.sheet_by_area.column_dimensions[col].width = value + 2

        # Бордеры первого листа
        for i in range(17):
            for j in range(5):
                self.sheet_by_year.cell(row=i + 1, column=j + 1).border = Border(top=thins, bottom=thins, left=thins,
                                                                                 right=thins)
        # Бордеры второго листа
        for i in range(11):
            for j in range(5):
                if j != 2:
                    self.sheet_by_area.cell(row=i + 1, column=j + 1).border = Border(top=thins, bottom=thins,
                                                                                     left=thins,
                                                                                     right=thins)
        # Формат ячейки для второго листа
        for i in range(10):
            self.sheet_by_area.cell(row=i + 2, column=5).number_format = "0.00%"


def parse_html(info):
    info = re.sub('<.*?>', '', info)
    info = info.replace("\r\n", "\n")
    res = [' '.join(word.split()) for word in info.split('\n')]
    return res[0] if len(res) == 1 else res


inputed = UserInput()
dataset = DataSet(inputed.file_name)
(names, all_vac_data) = dataset.names, dataset.all_data

static = StaticInfo()
for data in all_vac_data:
    parsed_data = Vacancy(dict(zip(names, map(parse_html, data))))
    parsed_data.get_vac_data(static, inputed.vacancy_name)

static.print_result(inputed.vacancy_name)

report = Report(inputed.vacancy_name, static.dicts_list_by_year, static.dicts_list_by_area, static.vac_with_others)
